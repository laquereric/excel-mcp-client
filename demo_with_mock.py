"""
Demo script with mock MCP data
Demonstrates how the plugin would work with actual MCP servers
"""

from mcp_connector import MCPServer, MCPCapability
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def create_demo_workbook():
    """Create a demo Excel workbook with mock MCP data"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MCPs"
    
    # Define colors
    COLOR_HEADER = "4472C4"  # Blue
    COLOR_SECTION = "D9D9D9"  # Light gray
    COLOR_CONNECTED = "C6E0B4"  # Light green
    COLOR_DISCONNECTED = "FFC7CE"  # Light red
    
    # Mock data - 3 example MCP servers
    mock_servers = [
        {
            "name": "weather-mcp",
            "status": "Connected",
            "tools": ["get_current_weather", "get_forecast", "get_historical_data"],
            "resources": ["weather://current", "weather://forecast/7day", "weather://stations"],
            "prompts": ["analyze_weather_pattern", "forecast_summary"]
        },
        {
            "name": "database-mcp",
            "status": "Connected",
            "tools": ["query_database", "execute_sql", "get_schema", "export_table"],
            "resources": ["db://schema", "db://tables", "db://views", "db://procedures"],
            "prompts": ["query_builder", "schema_analyzer", "optimization_suggestions"]
        },
        {
            "name": "filesystem-mcp",
            "status": "Disconnected",
            "tools": ["read_file", "write_file", "list_directory", "search_files"],
            "resources": ["file:///data/*", "file:///config/*", "file:///logs/*"],
            "prompts": ["file_summary", "directory_analysis"]
        }
    ]
    
    # Set up header row
    ws['A1'] = "Capability Type"
    ws['A1'].font = Font(bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="left", vertical="center")
    
    # Set column width
    ws.column_dimensions['A'].width = 25
    
    # Add server columns
    col = 2  # Start from column B
    for server_data in mock_servers:
        col_letter = openpyxl.utils.get_column_letter(col)
        
        # Server name header
        ws[f'{col_letter}1'] = server_data["name"]
        ws[f'{col_letter}1'].font = Font(bold=True, color="FFFFFF")
        ws[f'{col_letter}1'].fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type="solid")
        ws[f'{col_letter}1'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Status
        ws['A2'] = "Status"
        ws['A2'].font = Font(bold=True)
        ws[f'{col_letter}2'] = server_data["status"]
        
        if server_data["status"] == "Connected":
            ws[f'{col_letter}2'].fill = PatternFill(start_color=COLOR_CONNECTED, end_color=COLOR_CONNECTED, fill_type="solid")
        else:
            ws[f'{col_letter}2'].fill = PatternFill(start_color=COLOR_DISCONNECTED, end_color=COLOR_DISCONNECTED, fill_type="solid")
        
        # Set column width
        ws.column_dimensions[col_letter].width = 20
        
        col += 1
    
    # Build capability list (union of all capabilities)
    all_tools = set()
    all_resources = set()
    all_prompts = set()
    
    for server_data in mock_servers:
        all_tools.update(server_data["tools"])
        all_resources.update(server_data["resources"])
        all_prompts.update(server_data["prompts"])
    
    # Write capabilities
    row = 4  # Start from row 4
    
    # TOOLS section
    ws[f'A{row}'] = "TOOLS"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_SECTION, end_color=COLOR_SECTION, fill_type="solid")
    
    # Mark servers that have tools
    col = 2
    for server_data in mock_servers:
        col_letter = openpyxl.utils.get_column_letter(col)
        if server_data["tools"]:
            ws[f'{col_letter}{row}'] = "✓"
            ws[f'{col_letter}{row}'].font = Font(size=14)
            ws[f'{col_letter}{row}'].alignment = Alignment(horizontal="center")
        col += 1
    
    row += 1
    
    # Individual tools
    for tool in sorted(all_tools):
        ws[f'A{row}'] = tool
        
        # Mark which servers have this tool
        col = 2
        for server_data in mock_servers:
            col_letter = openpyxl.utils.get_column_letter(col)
            if tool in server_data["tools"]:
                ws[f'{col_letter}{row}'] = "✓"
                ws[f'{col_letter}{row}'].alignment = Alignment(horizontal="center")
            col += 1
        
        row += 1
    
    # Empty separator
    row += 1
    
    # RESOURCES section
    ws[f'A{row}'] = "RESOURCES"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_SECTION, end_color=COLOR_SECTION, fill_type="solid")
    
    # Mark servers that have resources
    col = 2
    for server_data in mock_servers:
        col_letter = openpyxl.utils.get_column_letter(col)
        if server_data["resources"]:
            ws[f'{col_letter}{row}'] = "✓"
            ws[f'{col_letter}{row}'].font = Font(size=14)
            ws[f'{col_letter}{row}'].alignment = Alignment(horizontal="center")
        col += 1
    
    row += 1
    
    # Individual resources
    for resource in sorted(all_resources):
        ws[f'A{row}'] = resource
        
        # Mark which servers have this resource
        col = 2
        for server_data in mock_servers:
            col_letter = openpyxl.utils.get_column_letter(col)
            if resource in server_data["resources"]:
                ws[f'{col_letter}{row}'] = "✓"
                ws[f'{col_letter}{row}'].alignment = Alignment(horizontal="center")
            col += 1
        
        row += 1
    
    # Empty separator
    row += 1
    
    # PROMPTS section
    ws[f'A{row}'] = "PROMPTS"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color=COLOR_SECTION, end_color=COLOR_SECTION, fill_type="solid")
    
    # Mark servers that have prompts
    col = 2
    for server_data in mock_servers:
        col_letter = openpyxl.utils.get_column_letter(col)
        if server_data["prompts"]:
            ws[f'{col_letter}{row}'] = "✓"
            ws[f'{col_letter}{row}'].font = Font(size=14)
            ws[f'{col_letter}{row}'].alignment = Alignment(horizontal="center")
        col += 1
    
    row += 1
    
    # Individual prompts
    for prompt in sorted(all_prompts):
        ws[f'A{row}'] = prompt
        
        # Mark which servers have this prompt
        col = 2
        for server_data in mock_servers:
            col_letter = openpyxl.utils.get_column_letter(col)
            if prompt in server_data["prompts"]:
                ws[f'{col_letter}{row}'] = "✓"
                ws[f'{col_letter}{row}'].alignment = Alignment(horizontal="center")
            col += 1
        
        row += 1
    
    # Freeze panes (first column and first two rows)
    ws.freeze_panes = "B3"
    
    # Save workbook
    output_path = "/home/ubuntu/excel_mcp_plugin/MCP_Demo.xlsx"
    wb.save(output_path)
    print(f"✅ Demo workbook created: {output_path}")
    print(f"\nThe workbook contains:")
    print(f"  - {len(mock_servers)} mock MCP servers")
    print(f"  - {len(all_tools)} tools")
    print(f"  - {len(all_resources)} resources")
    print(f"  - {len(all_prompts)} prompts")
    print(f"\nThis demonstrates the structure that the plugin creates.")
    
    return output_path


if __name__ == "__main__":
    print("=" * 60)
    print("Excel MCP Plugin - Demo with Mock Data")
    print("=" * 60)
    print()
    create_demo_workbook()
